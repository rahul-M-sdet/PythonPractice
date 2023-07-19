import openpyxl

book = openpyxl.load_workbook("C:\\Users\\rahul.m\\Desktop\\testdata.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A6'].value)
